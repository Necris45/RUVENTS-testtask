import openpyxl
import datetime


# Решето эратосфена для поиска простых чисел
def eratosfen_sieve(max_number):
    n = 2
    len_of_lst = max_number + 1
    sieve = [x for x in range(len_of_lst)]
    sieve[1] = 0
    while n < len_of_lst:
        if sieve[n] != 0:
            m = n * 2
            while m < len_of_lst:
                sieve[m] = 0
                m += n
        n += 1
    return [p for p in sieve if p != 0]


# открываем требуемый файл
wb = openpyxl.load_workbook('task_support.xlsx')
# назначаем активную страницу
wb.active = 1
sheet = wb.active
# определяем число строк
rows = sheet.max_row

# переменная для итогового значения первого задания
num1 = 0
# переменная для итогового значения второго задания
num2 = 0
# переменная для итогового значения третьего задания
num3 = 0
# переменная для итогового значения четвертого задания
date1 = 0
# переменная для итогового значения пятого задания
date2 = 0
# переменная для итогового значения шестого задания
date3 = 0

# Перебираем значения в требуемом столбце и считаем четные числа
for i in range(3, rows+1):
    cell = sheet.cell(row=i, column=2)
    if cell.value % 2 == 0:
        num1 += 1

# определяем максимум для решета эратосфена, чтобы создать список простых чисел в нужном диапазоне
max_value = 0
for i in range(3, rows+1):
    cell = sheet.cell(row=i, column=3)
    if cell.value > max_value:
        max_value = cell.value
# на основании максимума составляем список простых чисел не более максимального
simple_lst = eratosfen_sieve(max_value)
# сверяем все числа на предмет нахождения в списке и те, что есть в нем, считаем
for i in range(3, rows+1):
    cell = sheet.cell(row=i, column=3)
    if cell.value in simple_lst:
        num2 += 1

# в третьем задании работа со строками
for i in range(3, rows+1):
    cell = sheet.cell(row=i, column=4)
    # убираем из строк лишние пробелы
    value = cell.value.replace(" ", "")
    # заменяем "," на "."
    value = value.replace(",", ".")
    # если в начале не 0, то ставим сперва ноль, потом остальную строку
    if value[0] != '0':
        value = '0'+value
    # преобразуем строку в число с плавающей запятой
    value = float(value)
    # все что менее 0,5 считаем
    if value < 0.5:
        num3 += 1

# по прежнему работа со строкой. Так как в начале каждой строки 3 буквы обозначают день недели, просто считаем
for i in range(3, rows+1):
    cell = sheet.cell(row=i, column=5)
    # считаем
    if cell.value[0:3] == 'Tue':
        date1 += 1

# немногим сложнее. Преобразуем строку в дату, weekday() вернет день недели в виде числа. Вторник = 1.
for i in range(3, rows+1):
    cell = sheet.cell(row=i, column=6)
    # получаем строку с датой
    value = cell.value[0:10]
    # преобразуем строку в дату
    date = datetime.datetime.strptime(value, "%Y-%m-%d")
    # считаем
    if date.weekday() == 1:
        date2 += 1

# также не сильно сложнее, требуется всего одна дополнительная проверка
for i in range(3, rows+1):
    cell = sheet.cell(row=i, column=7)
    # получаем строку с датой
    value = cell.value
    # преобразуем строку в дату
    date = datetime.datetime.strptime(value, "%m-%d-%Y")
    # добавляем промежуточную дату
    date_2 = date + datetime.timedelta(days=7)
    # считаем
    if date.weekday() == 1 and (date.month != date_2.month):
        date3 += 1

print(f'В столбце num1 "{num1}" четных чисел')
print(f'В столбце num2 "{num2}" простых чисел')
print(f'В столбце num3 "{num3}" чисел меньше 0,5')
print(f'В столбце date1 "{date1}" вторников')
print(f'В столбце date2 "{date2}" вторников')
print(f'В столбце date3 "{date3}" последних вторников месяца')
